from openpyxl import load_workbook, Workbook

#open source sheet
path = 'C:/Users/family/Desktop/LA_1.xlsx'
wb = load_workbook(path)
ws1 = wb.worksheets[1]

wb.create_sheet('Dropped')
wb.active = 3
print(wb.active)
ws2 = wb.worksheets[3]
print(ws2)
# calculate total number of rows and
# columns in source excel sheet

max_r = ws1.max_row
max_c = ws1.max_column

# copying cell values from source
for i in range(1, max_r + 1):
    for j in range(1, max_c + 1):
        # reading cell value from source excel file
        c= ws1.cell(row=i, column= j)

        #writing the read value to destination sheet
        ws2.cell(row=i, column=j).value = c.value


del_rows = []
final_row = ws2.max_row

for i in (range(50)):
    if ws2.cell(i + 1, 16).value == None:
        del_rows.append(i + 1)
print(del_rows)

for row in reversed(del_rows):
    ws2.delete_rows(row)

wb.save(path)
