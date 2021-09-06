from typing import List, Any

from openpyxl import load_workbook, Workbook

class ContStudents:
    #open source sheet
    # path1 = 'C:/Users/family/Desktop/LA_1B.xlsx'
    path1 = 'LA_1B.xlsx'
    wb1 = load_workbook(path1)
    ws1 = wb1.worksheets[2]
    max_r = ws1.max_row

    def __init__(self):
        self.cont_students = []

    def cont_students_list(self):
        for i in range(2, ContStudents.max_r + 1):
            self.cont_students.append(ContStudents.ws1.cell(i, 2).value)
        print(self.cont_students)
        return self.cont_students

class NewStudents:

    path2 = 'C:/Users/family/Desktop/LA_2B.xlsx'
    wb2 = load_workbook(path2)
    ws2 = wb2.worksheets[0]
    max_r2 = ws2.max_row
    max_c2 = ws2.max_column

    def __init__(self, cont_students):
        self.cont_students = cont_students

    def new_students(self):
        for i in reversed(range(2, NewStudents.max_r2 + 1)):
            for num in reversed(self.cont_students):
                if num == NewStudents.ws2.cell(row=i, column=2).value:
                    row = i
                    NewStudents.ws2.delete_rows(int(row))
                    break
        NewStudents.wb2.save(NewStudents.path2)
        return NewStudents.ws2

class CopyNewStudents:
    path1 = 'LA_1B.xlsx'
    wb1 = load_workbook(path1)
    ws1 = wb1.worksheets[2]
    max_r = ws1.max_row-1

    def __init__(self, ws2):
        self.ws2 = ws2


    def copy_new_students(self):
        for i in range(2, NewStudents.max_r2 + 1):
            for j in range(2, NewStudents.max_c2 + 1):
                # reading cell value from source excel file
                c = self.ws2.cell(row=i, column=j)

                # writing the read value to destination sheet
                self.ws1.cell(row=i + CopyNewStudents.max_r, column=j).value = c.value
        return self.ws1
        # CopyNewStudents.wb1.save(CopyNewStudents.wb1)

c = ContStudents()
cont_students = c.cont_students_list()
n = NewStudents(cont_students=cont_students)
ws2 = n.new_students()
c2 = CopyNewStudents(ws2=ws2)
ws1 = c2.copy_new_students()
CopyNewStudents.wb1.save(CopyNewStudents.path1)




# max_r = ws2.max_row
# max_c = ws2.max_column
# print(max_r)
# # copying cell values from source
#
# max_r1 = ws1.max_row-1
# print(max_r1)
#
# for i in range(1, max_r + 1):
#     for j in range(1, max_c + 1):
#         # reading cell value from source excel file
#         c = ws2.cell(row=i, column=j)
#
#         # writing the read value to destination sheet
#         ws1.cell(row=i, column=j).value = c.value
#
# ws1.auto_filter.add_sort_condition(ref='B2:B67', descending=False)
# print(ws1)
# wb2.save(path2)